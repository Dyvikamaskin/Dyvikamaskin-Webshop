"""
Export the OEM DB's Machine + MachineRevision rows to an Excel file for
human review. Two sheets:

  Machines    — one row per Machine with key fields + revision counts
  Revisions   — one row per MachineRevision joined with Machine for context

Usage:
    python scripts/oem-utils/export-machines-xlsx.py
    # writes data/exports/oem-machines.xlsx
"""
import os
import re
import sys
from pathlib import Path

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[2]
ENV_FILE = REPO_ROOT / ".env"
OUT_PATH = REPO_ROOT / "data" / "exports" / "oem-machines.xlsx"


def load_env(path: Path) -> dict[str, str]:
    env = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        m = re.match(r'^([A-Z_][A-Z0-9_]*)=("?)(.*?)\2\s*$', line)
        if m:
            env[m.group(1)] = m.group(3)
    return env


def main():
    env = load_env(ENV_FILE)
    url = env.get("OEM_DIRECT_URL")
    if not url:
        raise SystemExit("OEM_DIRECT_URL not set in .env")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    conn = psycopg2.connect(url)
    cur = conn.cursor()

    # ─── Sheet 1: Machines ───
    cur.execute("""
        SELECT
            m.code,
            m.source::text,
            m."displayName",
            m."modelName",
            m."categoryPath"::text AS category_path,
            m."isDiscontinued",
            m."primaryImageUrl",
            (SELECT COUNT(*) FROM "MachineRevision" r WHERE r."machineId" = m.id) AS n_revisions,
            (SELECT COUNT(*) FROM "MachineRevision" r
             WHERE r."machineId" = m.id AND r."hasBom") AS n_revisions_with_bom,
            (SELECT COUNT(*) FROM "MachineRevision" r
             WHERE r."machineId" = m.id AND r."mode" = 'NUMERIC') AS n_numeric_revs,
            (SELECT COUNT(*) FROM "MachineRevision" r
             WHERE r."machineId" = m.id AND r."mode" = 'SERIAL_RANGE') AS n_serial_books,
            (SELECT COUNT(*) FROM "Diagram" d
             JOIN "MachineRevision" r ON r.id = d."revisionId"
             WHERE r."machineId" = m.id) AS n_diagrams,
            m.id
        FROM "Machine" m
        ORDER BY m."modelName", m.code
    """)
    machine_rows = cur.fetchall()
    print(f"Machines: {len(machine_rows):,}")

    # ─── Sheet 2: Revisions ───
    cur.execute("""
        SELECT
            m.code AS machine_code,
            m."modelName" AS model_name,
            m."displayName" AS machine_display_name,
            r."revisionTag",
            r."mode"::text,
            r."hasBom",
            r."sparePartListCode",
            r."afCode",
            r."aiCode",
            r."serialFrom",
            r."serialTo",
            r."rawName",
            r."partsManualFilename",
            r."partsManualUrl",
            (SELECT COUNT(*) FROM "Diagram" d WHERE d."revisionId" = r.id) AS n_diagrams,
            (CASE WHEN r."operatingManuals" IS NULL THEN 0
                  ELSE jsonb_array_length(r."operatingManuals") END) AS n_operating_manuals,
            m."categoryPath"::text AS category_path
        FROM "MachineRevision" r
        JOIN "Machine" m ON m.id = r."machineId"
        ORDER BY m."modelName", m.code, r."mode", r."revisionTag"
    """)
    revision_rows = cur.fetchall()
    print(f"Revisions: {len(revision_rows):,}")

    cur.close()
    conn.close()

    # ─── Write workbook ───
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D5F8F")

    ws1 = wb.active
    ws1.title = "Machines"
    m_headers = [
        "code", "source", "displayName", "modelName", "categoryPath",
        "isDiscontinued", "primaryImageUrl",
        "n_revisions", "n_revisions_with_bom", "n_numeric_revs",
        "n_serial_books", "n_diagrams", "id",
    ]
    ws1.append(m_headers)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.freeze_panes = "A2"
    for row in machine_rows:
        ws1.append(list(row))

    ws2 = wb.create_sheet("Revisions")
    r_headers = [
        "machine_code", "model_name", "machine_display_name",
        "revisionTag", "mode", "hasBom",
        "sparePartListCode", "afCode", "aiCode", "serialFrom", "serialTo",
        "rawName", "partsManualFilename", "partsManualUrl",
        "n_diagrams", "n_operating_manuals", "category_path",
    ]
    ws2.append(r_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.freeze_panes = "A2"
    for row in revision_rows:
        ws2.append(list(row))

    # Reasonable column widths
    width_hints = {
        "code": 14, "source": 18, "displayName": 36, "modelName": 16,
        "categoryPath": 50, "isDiscontinued": 10, "primaryImageUrl": 40,
        "n_revisions": 8, "n_revisions_with_bom": 12, "n_numeric_revs": 10,
        "n_serial_books": 10, "n_diagrams": 10, "id": 38,
        "machine_code": 14, "model_name": 16, "machine_display_name": 36,
        "revisionTag": 14, "mode": 14, "hasBom": 8, "sparePartListCode": 14,
        "afCode": 14, "aiCode": 14, "serialFrom": 22, "serialTo": 22,
        "rawName": 50, "partsManualFilename": 36, "partsManualUrl": 60,
        "n_operating_manuals": 12, "category_path": 50,
    }
    for ws in (ws1, ws2):
        for idx, h in enumerate(ws[1], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width_hints.get(h.value, 18)

    wb.save(OUT_PATH)
    print(f"\n✓ wrote {OUT_PATH}")
    print(f"   {len(machine_rows):,} machines, {len(revision_rows):,} revisions")


if __name__ == "__main__":
    main()
